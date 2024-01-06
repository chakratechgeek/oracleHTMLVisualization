"import requeired packages and class"
import urllib.request  as urllib2
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import xlsxwriter
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

"This class to extract SGA Target data and visualize it and store it in excel"
class SgaAdvisryTarget:

    sgaAdvisoryDF = pd.DataFrame()

    def __init__ (self):
        "Constructor"

    def drawTosheet(sheetName,pos, png):

        book = load_workbook('output.xlsx')
        newSheet=book.create_sheet(sheetName)
        newSheet.append(["SGA Target Advisory"])
        for r in dataframe_to_rows(SgaAdvisryTarget.sgaAdvisoryDF, index=False,header=True):
            newSheet.append(r)

        newSheet.add_image( openpyxl.drawing.image.Image(png),pos)
        book.save("output.xlsx")

    def sgatargetadvisory (self,soup):
        print ("\nExtracting SGA Advisory Target  details\n")
        cols = []
        rows=[]
        sgaAdvisoryTable = soup.find('table', summary='This table displays SGA target advisory for different SGA target sizes. It displays SGA size factor, estimated DB time and estimated physical reads for different SGA target sizes.')

        for row in sgaAdvisoryTable.find_all("tr"):
            cells = row.find_all('th')
            for val in cells:
                cols.append (val.find(text = True))

        count=0
        for row in sgaAdvisoryTable.find_all("tr"):
            cells = row.find_all('td')
            for val in cells:
                rows.append (val.find(text=True))

        x = [0, 4, 8, 12, 16,20,24,28,32,36,40,44,48,52,56]
        y = [4, 8, 12, 16,20,24,28,32,36,40,44,48,52,56,60]
        sgaAdvisoryDF = pd.DataFrame()
        length=len(rows)

        count=1
        for i, j in zip(x, y):
            SgaAdvisryTarget.sgaAdvisoryDF[count] = rows[i:j]
            count = count+1

        SgaAdvisryTarget.sgaAdvisoryDF = SgaAdvisryTarget.sgaAdvisoryDF.transpose()
        SgaAdvisryTarget.sgaAdvisoryDF.columns=cols
        SgaAdvisryTarget.sgaAdvisoryDF = SgaAdvisryTarget.sgaAdvisoryDF.iloc[1:]

        SgaAdvisryTarget.sgaAdvisoryDF.rename (columns = lambda x: x.strip().replace(' ','_'),inplace= True)
        SgaAdvisryTarget.sgaAdvisoryDF.rename (columns = lambda x: x.replace('___','_'),inplace= True)
        SgaAdvisryTarget.sgaAdvisoryDF.rename (columns = lambda x: x.replace('__','_'),inplace= True)

        for col in SgaAdvisryTarget.sgaAdvisoryDF.columns:
            SgaAdvisryTarget.sgaAdvisoryDF[col] = SgaAdvisryTarget.sgaAdvisoryDF[col].apply (lambda x: x.replace(',',''))

        for col in SgaAdvisryTarget.sgaAdvisoryDF.columns:
            SgaAdvisryTarget.sgaAdvisoryDF[col] = SgaAdvisryTarget.sgaAdvisoryDF[col].astype('float')

        plt.figure(figsize=(8, 5))
        plt.plot(SgaAdvisryTarget.sgaAdvisoryDF['SGA_Target_Size_(M)'],SgaAdvisryTarget.sgaAdvisoryDF['Est_DB_Time_(s)'])
        plt.plot(SgaAdvisryTarget.sgaAdvisoryDF['SGA_Target_Size_(M)'],SgaAdvisryTarget.sgaAdvisoryDF['Est_Physical_Reads'])
        plt.legend(['Est_DB_Time_(s)','Est_Physical_Reads'], prop={'size': 6})
        plt.title('SGA Target Advisory', size=13)
        plt.ylabel('Reads and Sec', size=8)
        plt.xlabel('SGA Target Size (M)', size=8)
        plt.xticks(size=5)
        plt.yticks(size=5)
        plt.savefig('sgaadvisory.png')
        SgaAdvisryTarget.drawTosheet('SGA_Advisory',"I1","sgaadvisory.png")