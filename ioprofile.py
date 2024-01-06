"import requeired packages and class"
import urllib.request  as urllib2
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup
import pandas as pd
import xlsxwriter
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

"This class to extract ADDM and Foreground wait event data and visualize it and store it in excel"
class IOProfile:

    def __init__ (self):
        "Constructor"

    def drawTosheet(sheetName,pos, png):

        book = load_workbook('output.xlsx')
        newSheet=book.create_sheet(sheetName)
        for r in dataframe_to_rows(IOProfile.ioDF, index=False,header=True):
            newSheet.append(r)

        for fig in png:
            newSheet.add_image( openpyxl.drawing.image.Image(fig),pos)
            pos="J10"
        book.save("output.xlsx")

    def ioprofile(self,soup):
        print ("\nExtracting IO Profile details\n")
        cols = []
        rows=[]
        ioTable = soup.find('table', summary='This table displays IO profile')

        for row in ioTable.find_all("tr"):
            cells = row.find_all('th')
            for val in cells:
                cols.append (val.find(text = True))

        for row in ioTable.find_all("tr"):
            cells = row.find_all('td')
            for val in cells:
                rows.append (val.find(text=True))

        x = [0, 4, 8, 12, 16,20,24,28,32,36,40]
        y = [4, 8, 12, 16,20,24,28,32,36,40,44]
        IOProfile.ioDF = pd.DataFrame(cols)
        count=1
        for i, j in zip(x, y):
            IOProfile.ioDF[count] = rows[i:j]
            count = count+1

        IOProfile.ioDF = IOProfile.ioDF.transpose()
        IOProfile.ioDF.columns=cols
        IOProfile.ioDF = IOProfile.ioDF.iloc[1:]
        IOProfile.ioDF.columns = IOProfile.ioDF.columns.fillna('Types')
        IOProfile.ioDF.rename (columns = lambda x: x.strip().replace(' ','_').lower(),inplace= True)

        for col in IOProfile.ioDF.columns:
            IOProfile.ioDF[col] = IOProfile.ioDF[col].apply (lambda x: x.replace(',',''))

        for col in IOProfile.ioDF.columns:
            if col != 'types':
                IOProfile.ioDF[col] = IOProfile.ioDF[col].astype('float')
        IOProfile.ioDF = IOProfile.ioDF.iloc[:8]
        plt.figure(figsize=(8, 5))
        plt.plot(IOProfile.ioDF['types'],IOProfile.ioDF['read+write_per_second'])
        plt.plot(IOProfile.ioDF['types'],IOProfile.ioDF['read_per_second'])
        plt.plot(IOProfile.ioDF['types'],IOProfile.ioDF['write_per_second'])
        plt.legend(['read+write_per_second','read_per_second','write_per_second'], prop={'size': 6})
        plt.title('IO Profile', size=15)
        plt.ylabel('# of IO Per Second', size=8)
        plt.xticks(rotation=5,size=5)
        plt.yticks(size=5)
        plt.savefig('ioprofile.png')
        IOProfile.drawTosheet('IO Profile','H2',["ioprofile.png"])