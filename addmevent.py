"import requeired packages and class"

import urllib.request  as urllib2

import matplotlib.pyplot as plt

from bs4 import BeautifulSoup

import pandas as pd

import xlsxwriter

from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

"This class to extract ADDM and Foreground wait event data and visualize it and store it in excel"
class AddmEvent:
    addmDF = pd.DataFrame()
    Foredf = pd.DataFrame()
    lenCols=0

    def __init__ (self):
        "Constructor"

    def drawTosheet(sheetName,pos, png):

        book = load_workbook('output.xlsx')
        newSheet=book.create_sheet(sheetName)

        if AddmEvent.lenCols != 0:
            newSheet.append(["ADDM DataFrame"])
            for r in dataframe_to_rows(AddmEvent.addmDF, index=False,header=True):
                newSheet.append(r)
        else:
            newSheet.append(["No ADDM DataFrame Found for this AWR report"])

        newSheet.append(["Top 10 Foreground Events by Total Wait Time"])
        for r in dataframe_to_rows(AddmEvent.Foredf, index=False,header=True):
            newSheet.append(r)

        count=0
        for fig in png:
            for val in range(count,len(pos)):
                position=pos[val]
                count=count+1
                newSheet.add_image( openpyxl.drawing.image.Image(fig),position)
                break
        book.save("output.xlsx")

    def Addm(self,soup):
        print ("\nExtracting ADDM and Foreground wait event details\n")
        cols = []
        rows=[]
        addmTable = soup.find('table', summary='This table displays top ADDM findings by average active sessions')
        if addmTable is None:
            AddmEvent.lenCols=0
        else:
            for row in addmTable.find_all("tr"):
                cells = row.find_all('th')
                for val in cells:
                    cols.append (val.find(text = True))
            AddmEvent.lenCols = int (len(cols))
            count=0
            for row in addmTable.find_all("tr"):
                if count < 7:
                    cells = row.find_all('td')
                    for val in cells:
                        rows.append (val.find(text=True))
                count=count+1

            x = [0, 6, 12, 18, 24]
            y = [6, 12, 18,24, 30]
            AddmEvent.addmDF = pd.DataFrame(cols)
            length=len(rows)
            count=1
            lengthRows = int(length/6)
            for i, j in zip(x, y):
                AddmEvent.addmDF[count] = rows[i:j]
                count = count+1
                if lengthRows < count:
                    break

            AddmEvent.addmDF = AddmEvent.addmDF.transpose()
            AddmEvent.addmDF.columns=cols
            AddmEvent.addmDF = AddmEvent.addmDF.iloc[1:]
            AddmEvent.addmDF = AddmEvent.addmDF.sort_values('Percent active sessions of finding',ascending =False)

            plt.figure(figsize=(5, 3))
            plt.bar(AddmEvent.addmDF['Finding Name'],AddmEvent.addmDF['Percent active sessions of finding'])
            plt.title('ADDM Report Summary', size=15)
            plt.xlabel('Events', size=5)
            plt.ylabel('Percent active sessions', size=5)
            plt.xticks(rotation= 10,size=5)
            plt.yticks(size=5)
            plt.savefig('addmbar.png')

            plt.figure(figsize=(7, 4))
            plt.title('ADDM Report Summary')
            colors = ['r', 'g','b','c','m']
            plt.pie(AddmEvent.addmDF['Percent active sessions of finding'], labels=AddmEvent.addmDF['Finding Name'], colors=colors, autopct='%.1f%%',shadow=True, startangle=90)
            plt.savefig('addmpie.png')

        "Call foreground wait event function"
        ForeTable = soup.find('table', summary='This table displays top 10 wait events by total wait time')
        AddmEvent.foreevent(ForeTable)

        "Call Image to Excel function"
        if AddmEvent.lenCols != 0:
            AddmEvent.drawTosheet('ADDM_ForeWait',["I1","K5","M10"],["addmpie.png","addmbar.png","fore.png"])
        else:
            AddmEvent.drawTosheet('ADDM_ForeWait',["I1"],["fore.png"])

    def foreevent(table):
        cols = []
        for row in table.find_all("tr"):
            cells = row.find_all('th')
            for val in cells:
                cols.append (val.find(text = True))

        A=[]
        B=[]
        C=[]
        D=[]
        E=[]
        F=[]

        for row in table.findAll("tr"):
            count=0
            cells = row.find_all('td')
            for val in cells:
                if count == 0:
                    A.append (val.find(text=True))
                elif  count == 1:
                    B.append (val.find(text=True))
                elif  count == 2:
                    C.append (val.find(text=True))
                elif  count == 3:
                    D.append (val.find(text=True))
                elif  count == 4:
                    E.append (val.find(text=True))
                elif  count == 5:
                    F.append (val.find(text=True))
                count= count+1

        AddmEvent.Foredf['Event']=A
        AddmEvent.Foredf['Waits']=B
        AddmEvent.Foredf['Total Wait Time (sec)']=C
        AddmEvent.Foredf['Avg Wait']=D
        AddmEvent.Foredf['% DB time']=E
        AddmEvent.Foredf['Wait Class']=F
        AddmEvent.Foredf.rename (columns = lambda x: x.strip().replace(' ','_'),inplace= True)
        AddmEvent.Foredf.rename (columns = lambda x: x.replace('__','_'),inplace= True)
        AddmEvent.Foredf.rename (columns = lambda x: x.replace('__','_'),inplace= True)
        AddmEvent.Foredf['Total_Wait_Time_(sec)'] = AddmEvent.Foredf['Total_Wait_Time_(sec)'].apply(lambda x: (float(x.replace('K', '')) * 1000) if x.find('K') != -1 else (x))
        AddmEvent.Foredf['%_DB_time'] = AddmEvent.Foredf['%_DB_time'].astype('float')
        AddmEvent.Foredf['Total_Wait_Time_(sec)'] = AddmEvent.Foredf['Total_Wait_Time_(sec)'].astype('float')
        AddmEvent.Foredf = AddmEvent.Foredf.sort_values('Total_Wait_Time_(sec)',ascending=False)

        plt.figure(figsize=(5, 3))
        plt.plot(AddmEvent.Foredf['Event'],AddmEvent.Foredf['Total_Wait_Time_(sec)'])
        plt.plot(AddmEvent.Foredf['Event'],AddmEvent.Foredf['%_DB_time'])
        plt.legend(['Total Wait Time(Sec)','DB time in %'], prop={'size': 8})
        plt.title('Top 10 Foreground Events', size=15)
        plt.xlabel('Events', size=5)
        plt.ylabel('Percent of DB Time and Wait event (sec)', size=5)
        plt.xticks(rotation= 10,size=5)
        plt.yticks(size=5)
        plt.savefig('fore.png')
