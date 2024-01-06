"import requeired packages and class"
import urllib.request  as urllib2
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import xlsxwriter
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl


"This class to extract SQL Statistics data and visualize it and store it in excel"
class SqlStats:

    def __init__ (self):
        "Constructor"

    def drawTosheet(sheetName,pos, png):

        book = load_workbook('output.xlsx')
        newSheet=book.create_sheet(sheetName)
        newSheet.append(["Sql Order By CPU"])
        for r in dataframe_to_rows(SqlStats.cpuDF, index=False,header=True):
            newSheet.append(r)

        newSheet.append(["Sql Order By Elapsed Time"])
        for r in dataframe_to_rows(SqlStats.esDF, index=False,header=True):
            newSheet.append(r)

        newSheet.append(["Sql Order By User I/O Wait"])
        for r in dataframe_to_rows(SqlStats.ioDF, index=False,header=True):
            newSheet.append(r)

        newSheet.append(["Sql Order By Gets"])
        for r in dataframe_to_rows(SqlStats.getsDF, index=False,header=True):
            newSheet.append(r)

        newSheet.append(["Sql Order By Physical Reads"])
        for r in dataframe_to_rows(SqlStats.prDF, index=False,header=True):
            newSheet.append(r)

        count=0
        for fig in png:
            for val in range(count,len(pos)):
                position=pos[val]
                count=count+1
                newSheet.add_image( openpyxl.drawing.image.Image(fig),position)
                break
        book.save("output.xlsx")

    def commonwrangle (table,cols):
        SqlStats.ssDF = pd.DataFrame()
        SqlStats.A=[]
        SqlStats.B=[]
        SqlStats.C=[]
        SqlStats.D=[]
        SqlStats.E=[]
        SqlStats.F=[]
        SqlStats.G=[]
        SqlStats.H=[]
        SqlStats.I=[]
        SqlStats.J=[]

        for row in table.findAll("tr"):
            count=0
            cells = row.find_all('td')
            for val in cells:
                if count == 0:
                    SqlStats.A.append (val.find(text=True))
                elif  count == 1:
                    SqlStats.B.append (val.find(text=True))
                elif  count == 2:
                    SqlStats.C.append (val.find(text=True))
                elif  count == 3:
                    SqlStats.D.append (val.find(text=True))
                elif  count == 4:
                    SqlStats.E.append (val.find(text=True))
                elif  count == 5:
                    SqlStats.F.append (val.find(text=True))
                elif  count == 6:
                    SqlStats.G.append (val.find(text=True))
                elif  count == 7:
                    SqlStats.H.append (val.find(text=True))
                elif  count == 8:
                    SqlStats.I.append (val.find(text=True))
                elif  count == 9:
                    SqlStats.J.append (val.find(text=True))
                count= count+1

            SqlStats.ssDF = pd.DataFrame(columns= cols)
            SqlStats.ssDF.rename (columns = lambda x: x.strip().replace(' ','_'),inplace= True)
            SqlStats.ssDF.rename (columns = lambda x: x.replace('__','_'),inplace= True)
            SqlStats.ssDF.rename (columns = lambda x: x.replace('__','_'),inplace= True)

            count=0
            alph = [SqlStats.A,SqlStats.B,SqlStats.C,SqlStats.D,SqlStats.E,SqlStats.F,SqlStats.G,SqlStats.H]
            for col in SqlStats.ssDF.columns:
                for val in range(count,len(alph)):
                    SqlStats.ssDF[col] = alph[count]
                    count= count+1
                    break

            colLength=len(SqlStats.ssDF.keys())
            SqlStats.ssDF=SqlStats.ssDF[SqlStats.ssDF.columns[: colLength-2]]

            for col in SqlStats.ssDF.columns:
                SqlStats.ssDF[col] = SqlStats.ssDF[col].apply (lambda x: x.replace(',',''))

            for col in SqlStats.ssDF.columns:
                SqlStats.ssDF[col] = SqlStats.ssDF[col].replace(r'^\s*$', np.nan, regex=True)
                SqlStats.ssDF[col].replace(np.nan,900,inplace=True)

            for col in SqlStats.ssDF.columns:
                if col != 'SQL_Id':
                    SqlStats.ssDF[col] = SqlStats.ssDF[col].astype('float')

    def cpustat(table):
        cols = []
        for row in table.find_all("tr"):
            cells = row.find_all('th')
            for val in cells:
                cols.append (val.find(text = True))
        SqlStats.commonwrangle (table,cols)

    def sqlstatistics(self,soup):
        print ("\nExtracting SQL Order By CPU details\n")
        cols = []
        rows=[]
        cpuTable=soup.find('table', summary='This table displays top SQL by CPU time')
        SqlStats.cpustat(cpuTable)
        SqlStats.cpuDF = SqlStats.ssDF
        SqlStats._cpuDF = SqlStats.ssDF
        SqlStats._cpuDF =SqlStats._cpuDF [(SqlStats._cpuDF ['%Total'] > 10) | (SqlStats._cpuDF['CPU_per_Exec_(s)'] > 0.1) | (SqlStats._cpuDF['CPU_Time_(s)'] > 30) | (SqlStats._cpuDF ['Elapsed_Time_(s)'] > 50) | (SqlStats._cpuDF['%CPU'] > 85)].sort_values('%Total',ascending=False).head(5)

        del SqlStats._cpuDF['Executions']
        ax=SqlStats._cpuDF.plot(kind='bar',figsize=(6,3))
        ax.set_xticklabels(SqlStats._cpuDF["SQL_Id"], rotation=0, fontsize=4)
        plt.title('Top 5 SQL Order By CPU', size=13)
        plt.ylabel('Percent and Sec', size=8)
        plt.xticks(rotation=10,size=5)
        plt.yticks(size=5)
        plt.savefig('orderbycpu.png')

    def elapsedsqlstatistics(self,soup):
        print ("\nExtracting SQL Order By Elapsed Time details\n")
        cols = []
        rows=[]
        elapsedTable=soup.find('table', summary='This table displays top SQL by elapsed time')
        SqlStats.cpustat(elapsedTable)
        SqlStats.esDF = SqlStats.ssDF
        SqlStats._esDF = SqlStats.ssDF
        SqlStats._esDF =SqlStats._esDF [(SqlStats._esDF ['%Total'] > 10) | (SqlStats._esDF['Elapsed_Time_per_Exec_(s)'] > 0.1)  | (SqlStats._esDF ['Elapsed_Time_(s)'] > 50) | (SqlStats._esDF['%CPU'] > 85)].sort_values('%Total',ascending=False).head(5)

        del SqlStats._esDF['Executions']
        ax=SqlStats._esDF.plot(kind='bar',figsize=(6,3))
        ax.set_xticklabels(SqlStats._esDF["SQL_Id"], rotation=0, fontsize=4)
        plt.title('Top 5 SQL Order By Elapsed Time', size=13)
        plt.ylabel('Percent and Sec', size=8)
        plt.xticks(rotation=10,size=5)
        plt.yticks(size=5)
        plt.savefig('orderbyelapsed.png')

    def iowaitsqlstatistics(self,soup):
        print ("\nExtracting SQL Order By USER IO details\n")
        cols = []
        rows=[]
        ioWaitTable=soup.find('table', summary='This table displays top SQL by user I/O time')
        SqlStats.cpustat(ioWaitTable)
        SqlStats.ioDF = SqlStats.ssDF
        SqlStats._ioDF = SqlStats.ssDF
        SqlStats._ioDF =SqlStats._ioDF [(SqlStats._ioDF ['%Total'] > 10) | (SqlStats._ioDF['UIO_per_Exec_(s)'] > 0.1)  | (SqlStats._ioDF ['Elapsed_Time_(s)'] > 50) | (SqlStats._ioDF['%CPU'] > 85)].sort_values('%Total',ascending=False).head(5)

        del SqlStats._ioDF['Executions']
        ax=SqlStats._ioDF.plot(kind='bar',figsize=(6,3))
        ax.set_xticklabels(SqlStats._ioDF["SQL_Id"], rotation=0, fontsize=4)
        plt.title('Top 5 SQL Order By USER IO', size=13)
        plt.ylabel('Percent and Sec', size=8)
        plt.xticks(rotation=10,size=5)
        plt.yticks(size=5)
        plt.savefig('orderbyio.png')


    def getssqlstatistics(self,soup):
        print ("\nExtracting SQL Order By Gets details\n")
        cols = []
        rows=[]
        getsTable=soup.find('table', summary='This table displays top SQL by buffer gets')
        SqlStats.cpustat(getsTable)
        SqlStats.getsDF = SqlStats.ssDF
        SqlStats._getsDF = SqlStats.ssDF
        SqlStats._getsDF =SqlStats._getsDF [(SqlStats._getsDF ['%Total'] > 10) | (SqlStats._getsDF['Gets__per_Exec'] > 0.1)  | (SqlStats._getsDF ['Elapsed_Time_(s)'] > 50) | (SqlStats._getsDF['%CPU'] > 85)].sort_values('%Total',ascending=False).head(5)

        del SqlStats._getsDF['Executions']
        ax=SqlStats._getsDF.plot(kind='bar',figsize=(6,3))
        ax.set_xticklabels(SqlStats._getsDF["SQL_Id"], rotation=0, fontsize=4)
        plt.title('Top 5 SQL Order By Gets', size=13)
        plt.ylabel('Percent and Sec', size=8)
        plt.xticks(rotation=10,size=5)
        plt.yticks(size=5)
        plt.savefig('orderbygets.png')

    def preadssqlstatistics(self,soup):
        print ("\nExtracting SQL Order By Physical Reads details\n")
        cols = []
        rows=[]
        pReadsTable=soup.find('table', summary='This table displays top SQL by physical reads')
        SqlStats.cpustat(pReadsTable)
        SqlStats.prDF = SqlStats.ssDF
        SqlStats._prDF = SqlStats.ssDF
        SqlStats._prDF =SqlStats._prDF [(SqlStats.prDF ['%Total'] > 10) | (SqlStats._prDF['Reads__per_Exec'] > 0.1)  | (SqlStats._prDF ['Elapsed_Time_(s)'] > 50) | (SqlStats._prDF['%CPU'] > 85)].sort_values('%Total',ascending=False).head(5)

        del SqlStats._prDF['Executions']
        ax=SqlStats._prDF.plot(kind='bar',figsize=(6,3))
        ax.set_xticklabels(SqlStats._prDF["SQL_Id"], rotation=0, fontsize=4)
        plt.title('Top 5 SQL Order By Physical Read', size=13)
        plt.ylabel('Percent and Sec', size=8)
        plt.xticks(rotation=10,size=5)
        plt.yticks(size=5)
        plt.savefig('orderbyPreads.png')
        SqlStats.drawTosheet("SQL Stats",["I1","K3","M5","Q6","I16"],["orderbycpu.png","orderbyelapsed.png","orderbyio.png","orderbygets.png","orderbyPreads.png"])
